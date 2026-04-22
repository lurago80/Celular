from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Count, Q, F
from django.http import HttpResponse, HttpResponseRedirect
from django.template.loader import render_to_string
from datetime import datetime, timedelta
from decimal import Decimal
import os

from .models import (
    Emitente, Cliente, Produto, Venda, ItemVenda,
    TradeIn, Compra, ItemCompra, Garantia, Licenca, OrdemServico, SerialProduto, Fornecedor, Devolucao
)
from .forms import VendaForm, ItemVendaFormSet, TradeInFormSet, ClienteForm, ProdutoForm, EmitenteForm, SerialForm, FornecedorForm, DevolucaoForm


def verificar_licenca_view(view_func):
    """Decorator para verificar se a licença está válida"""
    from functools import wraps
    
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # Verificar licença
        licenca = Licenca.licenca_valida()
        
        if not licenca:
            # Redirecionar para tela de ativação de licença
            return redirect('core:verificar_licenca')
        
        # Licença válida - continuar normalmente
        return view_func(request, *args, **kwargs)
    
    return _wrapped_view


import json

@login_required
@verificar_licenca_view
def dashboard(request):
    """Dashboard principal do sistema"""
    hoje = datetime.now().date()

    # Dados de vendas para o gráfico (últimos 14 dias)
    vendas_14d = []
    labels_14d = []
    for i in range(13, -1, -1):
        dia = hoje - timedelta(days=i)
        total = Venda.objects.filter(
            status='finalizada', data_venda__date=dia
        ).aggregate(t=Sum('total'))['t'] or 0
        labels_14d.append(dia.strftime('%d/%m'))
        vendas_14d.append(float(total))

    # Top 5 produtos mais vendidos (últimos 30 dias)
    top_produtos_qs = (
        ItemVenda.objects
        .filter(venda__status='finalizada', venda__data_venda__date__gte=hoje - timedelta(days=30))
        .values('produto__nome')
        .annotate(total_qtd=Sum('quantidade'))
        .order_by('-total_qtd')[:5]
    )
    top_labels = [p['produto__nome'] for p in top_produtos_qs]
    top_data = [p['total_qtd'] for p in top_produtos_qs]

    context = {
        'total_vendas': Venda.objects.filter(status='finalizada').count(),
        'vendas_abertas': Venda.objects.filter(status='aberta').count(),
        'total_clientes': Cliente.objects.count(),
        'total_produtos': Produto.objects.filter(ativo=True).count(),
        'produtos_estoque_baixo': Produto.objects.filter(ativo=True, estoque__lte=F('estoque_minimo')).count(),
        'vendas_hoje': Venda.objects.filter(
            status='finalizada',
            data_venda__date=hoje
        ).aggregate(Sum('total'))['total__sum'] or 0,
        'ultimas_vendas': Venda.objects.filter(status='aberta').order_by('-data_venda')[:5],
        # Ordens de Serviço
        'os_abertas': OrdemServico.objects.exclude(status__in=['entregue', 'cancelado']).count(),
        'os_aguardando': OrdemServico.objects.filter(status='aguardando').count(),
        'os_em_andamento': OrdemServico.objects.filter(status='em_andamento').count(),
        'os_concluidas': OrdemServico.objects.filter(status='concluido').count(),
        'ultimas_os': OrdemServico.objects.exclude(status__in=['entregue', 'cancelado']).order_by('-data_entrada')[:5],
        # Chart.js data (JSON strings)
        'chart_labels': json.dumps(labels_14d),
        'chart_vendas': json.dumps(vendas_14d),
        'chart_top_labels': json.dumps(top_labels),
        'chart_top_data': json.dumps(top_data),
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def nova_venda(request):
    """Criar nova venda"""
    if request.method == 'POST':
        form = VendaForm(request.POST)
        if form.is_valid():
            venda = form.save(commit=False)
            # data_venda é auto_now_add, então será preenchida automaticamente
            venda.save()
            
            formset = ItemVendaFormSet(request.POST, instance=venda)
            trade_formset = TradeInFormSet(request.POST, instance=venda)
            
            if formset.is_valid() and trade_formset.is_valid():
                # Verificar se há pelo menos um produto
                produtos_preenchidos = [f for f in formset.cleaned_data if f and f.get('produto')]
                if len(produtos_preenchidos) > 0:
                    itens_salvos = formset.save()
                    trade_formset.save()

                    # Vincular serial/IMEI selecionado a cada item
                    for idx, item in enumerate(itens_salvos):
                        serial_id = request.POST.get(f'serial_produto_{idx}', '').strip()
                        if serial_id:
                            try:
                                serial_obj = SerialProduto.objects.get(id=int(serial_id), status='em_estoque')
                                item.serial_produto = serial_obj
                                item.save(update_fields=['serial_produto'])
                            except (SerialProduto.DoesNotExist, ValueError):
                                pass
                    
                    venda.calcular_total()
                    messages.success(request, 'Venda criada com sucesso!')
                    return redirect('core:finalizar_venda', venda_id=venda.id)
                else:
                    messages.error(request, 'Adicione pelo menos um produto à venda.')
            else:
                messages.error(request, 'Erro ao processar itens da venda.')
        else:
            messages.error(request, 'Erro ao criar venda.')
    else:
        form = VendaForm()
        formset = ItemVendaFormSet()
        trade_formset = TradeInFormSet()
    
    context = {
        'form': form,
        'formset': formset,
        'trade_formset': trade_formset,
        'produtos': Produto.objects.filter(ativo=True).order_by('nome'),
    }
    return render(request, 'core/venda_nova.html', context)


@login_required
def finalizar_venda(request, venda_id):
    """Finalizar venda e gerar garantias"""
    venda = get_object_or_404(Venda, id=venda_id)
    
    if request.method == 'POST':
        # Validações ANTES de abrir transação
        emitente = Emitente.objects.first()
        if not emitente:
            messages.error(request, 'Emitente não configurado!')
            return redirect('admin:core_emitente_add')

        for item in venda.items.all():
            if not item.produto.tem_estoque(item.quantidade):
                messages.error(
                    request,
                    f'Estoque insuficiente para {item.produto.nome}. '
                    f'Disponível: {item.produto.estoque}'
                )
                return redirect('core:finalizar_venda', venda_id=venda_id)

            if item.produto.tipo == 'celular':
                # Se há serial vinculado, usar seus IMEIs (sem exigir POST)
                if item.serial_produto:
                    imei  = item.serial_produto.imei or ''
                    imei2 = item.serial_produto.imei2 or ''
                else:
                    # Verificar se foi selecionado serial na tela de finalização
                    serial_id_post = request.POST.get(f'serial_id_{item.id}', '').strip()
                    if serial_id_post:
                        try:
                            s = SerialProduto.objects.get(id=int(serial_id_post), status='em_estoque')
                            imei  = s.imei or ''
                            imei2 = s.imei2 or ''
                        except (SerialProduto.DoesNotExist, ValueError):
                            imei  = request.POST.get(f'imei_{item.id}', '').strip()
                            imei2 = request.POST.get(f'imei2_{item.id}', '').strip()
                    else:
                        imei  = request.POST.get(f'imei_{item.id}', '').strip()
                        imei2 = request.POST.get(f'imei2_{item.id}', '').strip()

                if not imei or len(imei) != 15 or not imei.isdigit():
                    messages.error(
                        request,
                        f'IMEI 1 obrigatório para {item.produto.nome} (deve ter exatamente 15 dígitos numéricos)'
                    )
                    return redirect('core:finalizar_venda', venda_id=venda_id)
                if not imei2 or len(imei2) != 15 or not imei2.isdigit():
                    messages.error(
                        request,
                        f'IMEI 2 obrigatório para {item.produto.nome} (deve ter exatamente 15 dígitos numéricos)'
                    )
                    return redirect('core:finalizar_venda', venda_id=venda_id)
                if imei == imei2:
                    messages.error(
                        request,
                        f'IMEI 1 e IMEI 2 de {item.produto.nome} não podem ser iguais'
                    )
                    return redirect('core:finalizar_venda', venda_id=venda_id)

        try:
            with transaction.atomic():
                # Finalizar venda
                venda.status = 'finalizada'
                venda.save()

                for item in venda.items.all():
                    # Vincular serial escolhido na tela de finalização (se ainda não vinculado)
                    if not item.serial_produto:
                        serial_id_post = request.POST.get(f'serial_id_{item.id}', '').strip()
                        if serial_id_post:
                            try:
                                serial_obj = SerialProduto.objects.get(id=int(serial_id_post), status='em_estoque')
                                item.serial_produto = serial_obj
                                item.save(update_fields=['serial_produto'])
                            except (SerialProduto.DoesNotExist, ValueError):
                                pass

                    # Baixar estoque
                    item.produto.estoque -= item.quantidade
                    item.produto.save()

                    # Calcular prazo da garantia
                    prazo = emitente.prazo_garantia_celular if item.produto.tipo == 'celular' else emitente.prazo_garantia_acessorio
                    data_inicio = datetime.now().date()
                    data_fim = data_inicio + timedelta(days=prazo)

                    # Determinar IMEIs: preferir serial vinculado, depois POST, depois produto
                    if item.serial_produto:
                        imei  = item.serial_produto.imei or ''
                        imei2 = item.serial_produto.imei2 or ''
                        # Marcar serial como vendido
                        from django.utils import timezone as tz
                        item.serial_produto.status = 'vendido'
                        item.serial_produto.data_saida = tz.now()
                        item.serial_produto.save(update_fields=['status', 'data_saida'])
                    else:
                        imei  = request.POST.get(f'imei_{item.id}', '')
                        imei2 = request.POST.get(f'imei2_{item.id}', '')

                    Garantia.objects.create(
                        venda=venda,
                        produto=item.produto,
                        item=item,
                        imei=imei if imei else None,
                        imei2=imei2 if imei2 else None,
                        data_inicio=data_inicio,
                        data_fim=data_fim,
                        prazo_dias=prazo,
                        texto=emitente.texto_garantia
                    )

            messages.success(request, f'Venda #{venda.id} finalizada com sucesso!')
            return redirect('core:venda_visualizar', venda_id=venda.id)

        except Exception as e:
            messages.error(request, f'Erro ao finalizar venda: {str(e)}')
    
    context = {
        'venda': venda,
        'itens': venda.items.select_related('serial_produto').all(),
        'tradeins': venda.tradeins.all(),
    }
    return render(request, 'core/venda_finalizar.html', context)


@login_required
def garantia_pdf(request, garantia_id):
    """Gerar HTML da garantia para impressão"""
    garantia = get_object_or_404(Garantia, id=garantia_id)
    emitente = Emitente.objects.first()
    
    if not emitente:
        messages.error(request, 'Emitente não configurado!')
        return redirect('core:dashboard')
    
    context = {
        'garantia': garantia,
        'emitente': emitente,
        'venda': garantia.venda,
        'cliente': garantia.venda.cliente,
    }
    
    return render(request, 'core/garantia_template.html', context)


@login_required
def venda_visualizar(request, venda_id):
    """Visualizar detalhes da venda"""
    venda = get_object_or_404(Venda, id=venda_id)
    
    context = {
        'venda': venda,
        'itens': venda.items.all(),
        'tradeins': venda.tradeins.all(),
    }
    return render(request, 'core/venda_visualizar.html', context)


@login_required
def comprovante_venda(request, venda_id):
    """Gerar comprovante de venda em HTML (para impressão)"""
    venda = get_object_or_404(Venda, id=venda_id)
    emitente = Emitente.objects.first()
    tipo = request.GET.get('tipo', 'a4')  # a4 ou cupom
    
    context = {
        'venda': venda,
        'emitente': emitente,
        'itens': venda.items.all(),
        'tradeins': venda.tradeins.all(),
        'tipo': tipo,
    }
    
    # Template baseado no tipo
    template_name = f'core/comprovante_{tipo}.html'
    
    return render(request, template_name, context)


@login_required
def termo_garantia(request, garantia_id):
    """Gerar HTML do termo de garantia para impressão"""
    garantia = get_object_or_404(Garantia, id=garantia_id)
    emitente = Emitente.objects.first()
    
    if not emitente:
        messages.error(request, 'Emitente não configurado!')
        return redirect('core:dashboard')
    
    # Permite ajustar o prazo de garantia antes de imprimir (sem alterar o banco)
    prazo_exibicao = garantia.prazo_dias
    data_fim_exibicao = garantia.data_fim
    prazo_param = request.GET.get('prazo_dias', '').strip()
    if prazo_param:
        try:
            prazo_custom = int(prazo_param)
            if prazo_custom > 0:
                prazo_exibicao = prazo_custom
                data_fim_exibicao = garantia.data_inicio + timedelta(days=prazo_custom)
        except (ValueError, TypeError):
            pass

    # Garantia Apple (informada manualmente antes de imprimir)
    garantia_apple = request.GET.get('garantia_apple', '').strip()

    context = {
        'garantia': garantia,
        'emitente': emitente,
        'venda': garantia.venda,
        'cliente': garantia.venda.cliente,
        'prazo_exibicao': prazo_exibicao,
        'data_fim_exibicao': data_fim_exibicao,
        'garantia_apple': garantia_apple,
    }
    
    return render(request, 'core/termo_garantia.html', context)


# ============ CADASTROS ============

@login_required
def cadastrar_cliente(request):
    """Cadastrar novo cliente com validação via Django Form"""
    form = ClienteForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            try:
                cliente = form.save()
                messages.success(request, f'Cliente {cliente.nome} cadastrado com sucesso!')
                return redirect('core:detalhe_cliente', cliente_id=cliente.id)
            except Exception as e:
                from django.db import IntegrityError
                if isinstance(e, IntegrityError) and 'unique' in str(e).lower():
                    messages.error(request, 'CPF ou e-mail já cadastrado para outro cliente.')
                else:
                    messages.error(request, f'Erro ao salvar cliente: {str(e)}')
        # erros do form são exibidos automaticamente no template
    return render(request, 'core/cadastro_cliente.html', {'form': form})


@login_required
def listar_clientes(request):
    """Listar clientes com busca e paginação"""
    qs = Cliente.objects.all().order_by('nome')
    busca = request.GET.get('busca', '').strip()
    if busca:
        qs = qs.filter(Q(nome__icontains=busca) | Q(cpf__icontains=busca) | Q(telefone__icontains=busca))
    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/lista_clientes.html', {'page_obj': page, 'busca': busca})


@login_required
def cadastrar_produto(request):
    """Cadastrar novo produto com validação via Django Form"""
    form = ProdutoForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            try:
                produto = form.save()
                messages.success(request, f'Produto {produto.nome} cadastrado com sucesso!')
                return redirect('core:listar_produtos')
            except Exception as e:
                messages.error(request, f'Erro ao salvar produto: {str(e)}')
    return render(request, 'core/cadastro_produto.html', {'form': form})


@login_required
def listar_produtos(request):
    """Listar produtos com filtros, alerta de estoque e paginação"""
    qs = Produto.objects.filter(ativo=True).select_related('fornecedor').order_by('nome')
    busca = request.GET.get('busca', '').strip()
    tipo = request.GET.get('tipo', '')
    estoque_critico = request.GET.get('critico', '')
    if busca:
        qs = qs.filter(Q(nome__icontains=busca) | Q(marca__icontains=busca) | Q(modelo__icontains=busca))
    if tipo:
        qs = qs.filter(tipo=tipo)
    if estoque_critico:
        qs = qs.filter(estoque__lte=F('estoque_minimo'))
    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/lista_produtos.html', {
        'page_obj': page,
        'busca': busca,
        'tipo': tipo,
        'estoque_critico': estoque_critico,
    })


@login_required
def cadastrar_usuario(request):
    """Cadastrar novo usuário do sistema"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email', '')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        
        # Validações
        if not username or not password:
            messages.error(request, 'Preencha os campos obrigatórios: Usuário e Senha.')
        elif password != password_confirm:
            messages.error(request, 'As senhas não coincidem!')
        elif User.objects.filter(username=username).exists():
            messages.error(request, f'Usuário "{username}" já existe!')
        elif email and User.objects.filter(email=email).exists():
            messages.error(request, f'E-mail "{email}" já está cadastrado!')
        else:
            try:
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_staff=is_staff,
                    is_superuser=is_superuser
                )
                messages.success(request, f'Usuário {username} cadastrado com sucesso!')
                return redirect('core:listar_usuarios')
            except Exception as e:
                messages.error(request, f'Erro ao cadastrar usuário: {str(e)}')
    
    return render(request, 'core/cadastro_usuario.html')


@login_required
def listar_usuarios(request):
    """Listar todos os usuários do sistema"""
    usuarios = User.objects.all().order_by('username')
    return render(request, 'core/lista_usuarios.html', {'usuarios': usuarios})


@login_required
def editar_usuario(request, usuario_id):
    """Editar dados de um usuário existente"""
    usuario = get_object_or_404(User, id=usuario_id)
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        email = request.POST.get('email', '')
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        nova_senha = request.POST.get('password', '').strip()
        confirmar_senha = request.POST.get('password_confirm', '').strip()

        if email and User.objects.filter(email=email).exclude(pk=usuario.pk).exists():
            messages.error(request, f'E-mail "{email}" já está cadastrado em outro usuário.')
        elif nova_senha and nova_senha != confirmar_senha:
            messages.error(request, 'As senhas não coincidem!')
        elif nova_senha and len(nova_senha) < 6:
            messages.error(request, 'A nova senha deve ter no mínimo 6 caracteres.')
        else:
            usuario.first_name = first_name
            usuario.last_name = last_name
            usuario.email = email
            usuario.is_staff = is_staff
            usuario.is_superuser = is_superuser
            usuario.is_active = is_active
            if nova_senha:
                usuario.set_password(nova_senha)
            usuario.save()
            messages.success(request, f'Usuário "{usuario.username}" atualizado com sucesso!')
            return redirect('core:listar_usuarios')

    return render(request, 'core/editar_usuario.html', {'usuario': usuario})


@login_required
def nova_compra(request):
    """Criar nova compra de fornecedor"""
    from .models import Fornecedor, Compra, ItemCompra
    
    if request.method == 'POST':
        try:
            # Validar fornecedor
            fornecedor_id = request.POST.get('fornecedor')
            if not fornecedor_id:
                messages.error(request, 'Selecione um fornecedor!')
                return render(request, 'core/nova_compra.html', {
                    'fornecedores': Fornecedor.objects.all(),
                    'produtos': Produto.objects.filter(ativo=True).order_by('nome'),
                })
            
            with transaction.atomic():
                # Criar compra
                compra = Compra.objects.create(
                    fornecedor_id=fornecedor_id,
                    total=0,
                    observacoes=request.POST.get('observacoes', ''),
                    forma_pagamento=request.POST.get('forma_pagamento', 'dinheiro')
                )

                # Processar itens
                produtos_ids = request.POST.getlist('produto')
                quantidades = request.POST.getlist('quantidade')
                precos = request.POST.getlist('preco_unitario')

                itens_criados = 0
                for produto_id, qtd, preco in zip(produtos_ids, quantidades, precos):
                    if produto_id and qtd and preco:
                        ItemCompra.objects.create(
                            compra=compra,
                            produto_id=produto_id,
                            quantidade=int(qtd),
                            preco_unitario=Decimal(str(preco))
                        )
                        itens_criados += 1

                if itens_criados == 0:
                    raise ValueError('Adicione pelo menos um item à compra!')

            messages.success(request, f'Compra #{compra.id} registrada com {itens_criados} item(ns)! Estoque atualizado.')
            return redirect('core:recibo_compra', compra_id=compra.id)

        except ValueError as e:
            messages.error(request, str(e))
            return render(request, 'core/nova_compra.html', {
                'fornecedores': Fornecedor.objects.all(),
                'produtos': Produto.objects.filter(ativo=True).order_by('nome'),
            })
        except Exception as e:
            messages.error(request, f'Erro ao registrar compra: {str(e)}')
    
    fornecedores = Fornecedor.objects.all()
    produtos = Produto.objects.filter(ativo=True).order_by('nome')
    
    context = {
        'fornecedores': fornecedores,
        'produtos': produtos,
    }
    
    return render(request, 'core/nova_compra.html', context)


@login_required
def listar_compras(request):
    """Listar compras com filtros, select_related e paginação"""
    from .models import Compra, Fornecedor

    qs = Compra.objects.select_related('fornecedor').order_by('-data_compra')

    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    fornecedor_id = request.GET.get('fornecedor')

    if data_inicio:
        qs = qs.filter(data_compra__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data_compra__lte=data_fim)
    if fornecedor_id:
        qs = qs.filter(fornecedor_id=fornecedor_id)

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page,
        'fornecedores': Fornecedor.objects.all().order_by('nome'),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'fornecedor_id': fornecedor_id,
    }

    return render(request, 'core/lista_compras.html', context)


@login_required
def recibo_compra(request, compra_id):
    """Gerar recibo de compra para o fornecedor"""
    from .models import Compra
    
    compra = Compra.objects.get(id=compra_id)
    emitente = Emitente.objects.first()
    
    context = {
        'compra': compra,
        'emitente': emitente,
    }
    
    return render(request, 'core/recibo_compra.html', context)


@login_required
def configurar_emitente(request):
    """Lista emitentes/filiais e permite criar/editar via EmitenteForm"""
    emitentes = Emitente.objects.all().order_by('nome')
    # editar emitente existente
    edit_id = request.GET.get('editar')
    emitente_edit = get_object_or_404(Emitente, id=edit_id) if edit_id else None

    form = EmitenteForm(
        request.POST or None,
        instance=emitente_edit,
    )
    if request.method == 'POST':
        if form.is_valid():
            obj = form.save(commit=False)
            if not Emitente.objects.exists():
                obj.ativo = True  # primeiro emitente fica ativo por padrão
            obj.save()
            messages.success(request, 'Emitente salvo com sucesso!')
            return redirect('core:configurar_emitente')

    return render(request, 'core/config_emitente.html', {
        'emitentes': emitentes,
        'form': form,
        'emitente_edit': emitente_edit,
    })


@login_required
def emitente_set_ativo(request, emitente_id):
    """Define o emitente selecionado como ativo"""
    emitente = get_object_or_404(Emitente, id=emitente_id)
    emitente.set_ativo()
    messages.success(request, f'{emitente.nome} definido como emitente ativo.')
    return redirect('core:configurar_emitente')


@login_required
def emitente_deletar(request, emitente_id):
    """Remove emitente que não seja o único"""
    emitente = get_object_or_404(Emitente, id=emitente_id)
    if Emitente.objects.count() <= 1:
        messages.error(request, 'Não é possível remover o único emitente.')
        return redirect('core:configurar_emitente')
    if request.method == 'POST':
        emitente.delete()
        messages.success(request, 'Emitente removido.')
    return redirect('core:configurar_emitente')


@login_required
def pdv(request):
    """Tela de PDV (Ponto de Venda) - Frente de Caixa"""
    from .models import Cliente
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Atualizar Cliente
        if action == 'atualizar_cliente':
            venda_id = request.session.get('venda_atual_id')
            if venda_id:
                venda = Venda.objects.get(id=venda_id)
                cliente_id = request.POST.get('cliente_id')
                venda.cliente_id = cliente_id
                venda.save()
                messages.success(request, 'Cliente atualizado!')
            return redirect('core:pdv')
        
        # Atualizar Vendedor
        if action == 'atualizar_vendedor':
            venda_id = request.session.get('venda_atual_id')
            if venda_id:
                venda = Venda.objects.get(id=venda_id)
                vendedor = request.POST.get('vendedor')
                venda.vendedor = vendedor
                venda.save()
                messages.success(request, 'Vendedor atualizado!')
            return redirect('core:pdv')
        
        # Processar adição de produto
        produto_id = request.POST.get('produto_id')
        try:
            quantidade = int(float(request.POST.get('quantidade', 1)))  # Converter para int
        except (ValueError, TypeError):
            quantidade = 0
        
        preco_unitario = request.POST.get('preco_unitario')  # Valor personalizado
        
        if produto_id and quantidade > 0:
            try:
                produto = Produto.objects.get(id=produto_id, ativo=True)
                
                # Verificar estoque
                if not produto.tem_estoque(quantidade):
                    messages.error(request, f'Estoque insuficiente. Disponível: {produto.estoque}')
                else:
                    # Criar ou obter venda atual (usar session)
                    venda_id = request.session.get('venda_atual_id')
                    if venda_id:
                        venda = Venda.objects.get(id=venda_id)
                    else:
                        # Criar nova venda com "CONSUMIDOR FINAL"
                        cliente_consumidor_final = Cliente.objects.filter(cpf__icontains='00000000000').first()
                        if not cliente_consumidor_final:
                            # Criar cliente "CONSUMIDOR FINAL"
                            cliente_consumidor_final = Cliente.objects.create(
                                nome='CONSUMIDOR FINAL',
                                cpf='00000000000',
                                telefone='',
                                email='consumidor@final.com',
                                endereco=''
                            )
                        venda = Venda.objects.create(
                            cliente=cliente_consumidor_final,
                            vendedor=request.user.username,
                            status='aberta'
                        )
                        request.session['venda_atual_id'] = venda.id
                    
                    # Usar preço personalizado se informado, senão usar preço do produto
                    try:
                        preco_final = Decimal(str(preco_unitario).replace(',', '.')) if preco_unitario else produto.preco_venda
                    except:
                        preco_final = produto.preco_venda
                    
                    # Verificar se produto já existe na venda
                    item_existente = ItemVenda.objects.filter(venda=venda, produto=produto).first()
                    if item_existente:
                        item_existente.quantidade += quantidade
                        # Recalcular subtotal
                        item_existente.subtotal = Decimal(str(item_existente.preco_unitario)) * item_existente.quantidade
                        item_existente.save()
                    else:
                        ItemVenda.objects.create(
                            venda=venda,
                            produto=produto,
                            quantidade=quantidade,
                            preco_unitario=preco_final
                        )
                    
                    venda.calcular_total()
                    messages.success(request, f'{produto.nome} adicionado!')
            except Exception as e:
                messages.error(request, f'Erro ao adicionar produto: {str(e)}')
        elif quantidade <= 0:
            messages.error(request, 'Quantidade deve ser maior que zero!')
    
    # Obter venda atual da sessão
    venda_id = request.session.get('venda_atual_id')
    if venda_id:
        try:
            venda = Venda.objects.get(id=venda_id, status='aberta')
        except:
            venda = None
            request.session.pop('venda_atual_id', None)
    else:
        venda = None
    
    # Buscar ou criar "CONSUMIDOR FINAL"
    consumidor_final = Cliente.objects.filter(nome='CONSUMIDOR FINAL').first()
    if not consumidor_final:
        consumidor_final = Cliente.objects.create(
            nome='CONSUMIDOR FINAL',
            cpf='00000000000',
            telefone='',
            email='consumidor@final.com',
            endereco=''
        )
    
    # Buscar clientes excluindo CONSUMIDOR FINAL para não duplicar
    clientes_ordenados = [consumidor_final] + list(Cliente.objects.exclude(nome='CONSUMIDOR FINAL').order_by('nome'))
    
    context = {
        'produtos': Produto.objects.filter(ativo=True).order_by('nome'),
        'venda': venda,
        'itens': venda.items.all() if venda else [],
        'clientes': clientes_ordenados,
    }
    return render(request, 'core/pdv.html', context)


@login_required
def pdv_limpar(request):
    """Limpar venda atual"""
    request.session.pop('venda_atual_id', None)
    messages.success(request, 'Carrinho limpo!')
    return redirect('core:pdv')


@login_required
def pdv_finalizar(request):
    """Finalizar venda do PDV"""
    venda_id = request.session.get('venda_atual_id')
    if not venda_id:
        messages.error(request, 'Nenhuma venda em andamento.')
        return redirect('core:pdv')
    
    try:
        venda = Venda.objects.get(id=venda_id, status='aberta')
        
        # Validar se há itens na venda
        if not venda.items.exists():
            messages.error(request, 'Adicione pelo menos um produto antes de finalizar!')
            return redirect('core:pdv')
        
        # Salvar desconto e forma de pagamento
        desconto_valor = request.GET.get('desconto', '0')
        forma_pagamento = request.GET.get('pagamento', 'dinheiro')
        
        try:
            venda.desconto = Decimal(str(desconto_valor))
        except:
            venda.desconto = Decimal('0')
        
        venda.desconto_percentual = Decimal('0')
        venda.forma_pagamento = forma_pagamento
        venda.calcular_total()
        
        # Validar estoque
        for item in venda.items.all():
            if not item.produto.tem_estoque(item.quantidade):
                messages.error(request, f'Estoque insuficiente para {item.produto.nome}')
                return redirect('core:pdv')
        
        with transaction.atomic():
            venda.save()

        # Redirecionar para tela de finalização
        request.session.pop('venda_atual_id', None)
        return redirect('core:finalizar_venda', venda_id=venda.id)

    except Venda.DoesNotExist:
        messages.error(request, 'Venda não encontrada.')
        request.session.pop('venda_atual_id', None)
        return redirect('core:pdv')
    except Exception as e:
        messages.error(request, f'Erro ao finalizar: {str(e)}')
        return redirect('core:pdv')


# ============ CONFERÊNCIA DE CAIXA ============

@login_required
def caixa_abrir(request):
    """Abrir novo caixa"""
    from .models import Caixa
    from datetime import datetime
    
    if request.method == 'POST':
        valor_inicial = request.POST.get('valor_inicial', 0)
        observacoes = request.POST.get('observacoes', '')
        
        # Verificar se há caixa aberto
        caixa_aberto = Caixa.objects.filter(status='aberto').first()
        if caixa_aberto:
            messages.warning(request, 'Já existe um caixa aberto! Feche-o antes de abrir outro.')
            return redirect('core:caixa_conferir', caixa_id=caixa_aberto.id)
        
        # Criar novo caixa
        caixa = Caixa.objects.create(
            operador=request.user.username,
            valor_inicial=valor_inicial,
            observacoes=observacoes
        )
        messages.success(request, f'Caixa #{caixa.id} aberto com sucesso!')
        return redirect('core:caixa_conferir', caixa_id=caixa.id)
    
    return render(request, 'core/caixa_abrir.html')


@login_required
def caixa_conferir(request, caixa_id):
    """Conferir caixa atual"""
    from .models import Caixa
    
    caixa = get_object_or_404(Caixa, id=caixa_id)
    
    if request.method == 'POST':
        # Atualizar contagem física
        caixa.notas_100 = int(request.POST.get('notas_100', 0))
        caixa.notas_50 = int(request.POST.get('notas_50', 0))
        caixa.notas_20 = int(request.POST.get('notas_20', 0))
        caixa.notas_10 = int(request.POST.get('notas_10', 0))
        caixa.notas_5 = int(request.POST.get('notas_5', 0))
        caixa.notas_2 = int(request.POST.get('notas_2', 0))
        caixa.moedas = Decimal(str(request.POST.get('moedas', 0)))
        caixa.observacoes = request.POST.get('observacoes', '')
        caixa.save()
        
        if request.POST.get('fechar') == 'true':
            return redirect('core:caixa_fechar', caixa_id=caixa.id)
        else:
            messages.success(request, 'Contagem atualizada!')
    
    # Obter vendas do período
    vendas = caixa.vendas_periodo()
    total_vendas = sum(v.total for v in vendas)
    
    # Calcular por forma de pagamento
    vendas_dinheiro = vendas.filter(forma_pagamento='dinheiro')
    vendas_debito = vendas.filter(forma_pagamento='debito')
    vendas_credito = vendas.filter(forma_pagamento='credito')
    vendas_pix = vendas.filter(forma_pagamento='pix')
    vendas_parcelado = vendas.filter(forma_pagamento='parcelado')
    
    total_dinheiro = sum(v.total for v in vendas_dinheiro)
    total_debito = sum(v.total for v in vendas_debito)
    total_credito = sum(v.total for v in vendas_credito)
    total_pix = sum(v.total for v in vendas_pix)
    total_parcelado = sum(v.total for v in vendas_parcelado)
    
    # Contagem física
    total_fisico = caixa.calcular_contagem_fisica()
    
    context = {
        'caixa': caixa,
        'vendas': vendas,
        'total_vendas': total_vendas,
        'total_dinheiro': total_dinheiro,
        'total_debito': total_debito,
        'total_credito': total_credito,
        'total_pix': total_pix,
        'total_parcelado': total_parcelado,
        'total_fisico': total_fisico,
        'valor_esperado': Decimal(str(caixa.valor_inicial)) + total_dinheiro,
        'diferenca': total_fisico - (Decimal(str(caixa.valor_inicial)) + total_dinheiro),
    }
    
    return render(request, 'core/caixa_conferir.html', context)


@login_required
def caixa_fechar(request, caixa_id):
    """Fechar caixa"""
    from .models import Caixa
    from datetime import datetime
    
    caixa = get_object_or_404(Caixa, id=caixa_id)
    
    if caixa.status == 'fechado':
        messages.warning(request, 'Este caixa já está fechado!')
        return redirect('core:caixa_historico')
    
    if request.method == 'POST':
        caixa.data_fechamento = datetime.now()
        caixa.valor_final = caixa.calcular_contagem_fisica()
        caixa.status = 'fechado'
        caixa.save()
        
        messages.success(request, f'Caixa #{caixa.id} fechado com sucesso!')
        return redirect('core:caixa_historico')
    
    return redirect('core:caixa_conferir', caixa_id=caixa.id)


@login_required
def caixa_historico(request):
    """Histórico de caixas"""
    from .models import Caixa
    
    caixas = Caixa.objects.all().order_by('-data_abertura')
    
    context = {
        'caixas': caixas,
    }
    
    return render(request, 'core/caixa_historico.html', context)


# ============ RELATÓRIOS ============

@login_required
def relatorio_produtos_comprados(request):
    """Relatório de produtos comprados"""
    from .models import ItemCompra
    
    itens = ItemCompra.objects.all().order_by('-compra__data_compra')
    
    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    if data_inicio:
        itens = itens.filter(compra__data_compra__gte=data_inicio)
    if data_fim:
        itens = itens.filter(compra__data_compra__lte=data_fim)
    
    context = {
        'itens': itens,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'total_compras': sum(item.subtotal for item in itens),
        'emitente': Emitente.get_ativo(),
    }
    
    return render(request, 'core/relatorios/produtos_comprados.html', context)


@login_required
def relatorio_produtos_vendidos(request):
    """Relatório de produtos vendidos"""
    itens = ItemVenda.objects.filter(venda__status='finalizada').select_related('venda__cliente', 'produto').order_by('-venda__data_venda')

    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    produto_id = request.GET.get('produto')

    if data_inicio:
        itens = itens.filter(venda__data_venda__gte=data_inicio)
    if data_fim:
        itens = itens.filter(venda__data_venda__lte=data_fim)
    if produto_id:
        itens = itens.filter(produto_id=produto_id)

    totais = itens.aggregate(total_vendas=Sum('subtotal'), total_quantidade=Sum('quantidade'))

    context = {
        'itens': itens,
        'produtos': Produto.objects.filter(ativo=True).order_by('nome'),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'produto_id': produto_id,
        'total_vendas': totais['total_vendas'] or 0,
        'total_quantidade': totais['total_quantidade'] or 0,
        'emitente': Emitente.get_ativo(),
    }

    return render(request, 'core/relatorios/produtos_vendidos.html', context)


@login_required
def relatorio_clientes(request):
    """Relatório de listagem de clientes"""
    clientes = Cliente.objects.all().order_by('nome')
    
    # Busca por nome
    busca = request.GET.get('busca', '')
    if busca:
        clientes = clientes.filter(nome__icontains=busca) | clientes.filter(cpf__icontains=busca)
    
    context = {
        'clientes': clientes,
        'busca': busca,
        'emitente': Emitente.get_ativo(),
    }
    
    return render(request, 'core/relatorios/clientes.html', context)


@login_required
def relatorio_vendas_periodo(request):
    """Relatório de vendas por período"""
    vendas = Venda.objects.filter(status='finalizada').select_related('cliente').order_by('-data_venda')

    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    cliente_id = request.GET.get('cliente')

    if data_inicio:
        vendas = vendas.filter(data_venda__gte=data_inicio)
    if data_fim:
        vendas = vendas.filter(data_venda__lte=data_fim)
    if cliente_id:
        vendas = vendas.filter(cliente_id=cliente_id)

    # Estatísticas usando aggregate (evita N+1)
    agg = vendas.aggregate(total_receita=Sum('total'), total_vendas=Count('id'))
    total_vendas = agg['total_vendas'] or 0
    total_receita = agg['total_receita'] or Decimal('0')
    ticket_medio = total_receita / total_vendas if total_vendas > 0 else 0

    context = {
        'vendas': vendas,
        'clientes': Cliente.objects.all().order_by('nome'),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'cliente_id': cliente_id,
        'total_vendas': total_vendas,
        'total_receita': total_receita,
        'ticket_medio': ticket_medio,
        'emitente': Emitente.get_ativo(),
    }

    return render(request, 'core/relatorios/vendas_periodo.html', context)


@login_required
def relatorio_curva_abc(request):
    """Relatório Curva ABC de produtos"""
    from django.db.models import Sum
    
    # Agrupar produtos vendidos
    produtos_vendidos = Produto.objects.filter(
        items__venda__status='finalizada'
    ).annotate(
        total_vendido=Sum('items__subtotal'),
        total_quantidade=Sum('items__quantidade')
    ).order_by('-total_vendido')
    
    # Calcular totais
    total_geral = sum(p.total_vendido or 0 for p in produtos_vendidos)
    
    # Classificar em A, B, C
    acumulado = 0
    for produto in produtos_vendidos:
        if total_geral > 0:
            porcentagem = (produto.total_vendido or 0) / total_geral * 100
        else:
            porcentagem = 0
        acumulado += porcentagem
        
        if acumulado <= 80:
            produto.classe = 'A'
        elif acumulado <= 95:
            produto.classe = 'B'
        else:
            produto.classe = 'C'
        
        produto.porcentagem = porcentagem
        produto.acumulado = acumulado
    
    context = {
        'produtos': produtos_vendidos,
        'total_geral': total_geral,
        'emitente': Emitente.get_ativo(),
    }
    
    return render(request, 'core/relatorios/curva_abc.html', context)


@login_required
def relatorio_inventario(request):
    """Relatório de inventário - estoque"""
    produtos = Produto.objects.filter(ativo=True).order_by('nome')
    
    # Filtros
    tipo = request.GET.get('tipo')
    busca = request.GET.get('busca', '')
    
    if tipo:
        produtos = produtos.filter(tipo=tipo)
    if busca:
        produtos = produtos.filter(nome__icontains=busca) | produtos.filter(marca__icontains=busca)
    
    # Estatísticas
    total_produtos = produtos.count()
    estoque_total = sum(p.estoque for p in produtos)
    valor_total = sum(p.preco_custo * p.estoque for p in produtos)
    produtos_zerados = produtos.filter(estoque=0).count()
    
    context = {
        'produtos': produtos,
        'total_produtos': total_produtos,
        'estoque_total': estoque_total,
        'valor_total': valor_total,
        'produtos_zerados': produtos_zerados,
        'tipo': tipo,
        'busca': busca,
        'emitente': Emitente.get_ativo(),
    }
    
    return render(request, 'core/relatorios/inventario.html', context)


# ============ BACKUP E RESTORE ============

@login_required
def backup_restore(request):
    """Página principal de backup e restore"""
    from django.conf import settings
    import os
    
    backups_dir = os.path.join(settings.BASE_DIR, 'backups')
    backups = []
    
    if os.path.exists(backups_dir):
        for file in sorted(os.listdir(backups_dir), reverse=True):
            if file.endswith('.db'):
                file_path = os.path.join(backups_dir, file)
                file_size = os.path.getsize(file_path)
                file_date = datetime.fromtimestamp(os.path.getmtime(file_path))
                backups.append({
                    'name': file,
                    'path': file_path,
                    'size': file_size,
                    'date': file_date
                })
    
    context = {
        'backups': backups,
    }
    
    return render(request, 'core/backup_restore.html', context)


@login_required
def backup_criar(request):
    """Criar backup do banco de dados"""
    from django.conf import settings
    import os
    import shutil
    
    try:
        # Criar diretório de backups se não existir
        backups_dir = os.path.join(settings.BASE_DIR, 'backups')
        os.makedirs(backups_dir, exist_ok=True)
        
        # Nome do arquivo de backup
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'backup_{timestamp}.db'
        backup_path = os.path.join(backups_dir, backup_filename)
        
        # Fazer cópia do banco de dados
        db_path = settings.DATABASES['default']['NAME']
        shutil.copy2(db_path, backup_path)
        
        # Calcular tamanho
        size_kb = os.path.getsize(backup_path) / 1024
        
        messages.success(request, f'Backup criado com sucesso! Arquivo: {backup_filename} ({(size_kb):.2f} KB)')
        
    except Exception as e:
        messages.error(request, f'Erro ao criar backup: {str(e)}')
    
    return redirect('core:backup_restore')


@login_required
def backup_restaurar(request):
    """Restaurar backup do banco de dados"""
    from django.conf import settings
    import os
    import shutil
    
    if request.method == 'POST':
        # Aceita APENAS o nome do arquivo, nunca um caminho completo (evita Path Traversal)
        backup_name = request.POST.get('backup_file', '').strip()
        
        if not backup_name:
            messages.error(request, 'Selecione um arquivo de backup!')
            return redirect('core:backup_restore')

        # Sanidade: rejeitar qualquer separador de diretório no nome
        if os.sep in backup_name or '/' in backup_name or '..' in backup_name:
            messages.error(request, 'Nome de arquivo inválido!')
            return redirect('core:backup_restore')

        backups_dir = os.path.join(settings.BASE_DIR, 'backups')
        backup_path = os.path.join(backups_dir, backup_name)

        # Confirmar que o caminho resolvido está dentro do diretório permitido
        backup_path = os.path.realpath(backup_path)
        backups_dir_real = os.path.realpath(backups_dir)
        if not backup_path.startswith(backups_dir_real + os.sep):
            messages.error(request, 'Acesso negado ao arquivo solicitado!')
            return redirect('core:backup_restore')

        if not os.path.exists(backup_path):
            messages.error(request, 'Arquivo de backup não encontrado!')
            return redirect('core:backup_restore')

        if not backup_name.endswith('.db'):
            messages.error(request, 'Formato de arquivo inválido!')
            return redirect('core:backup_restore')

        try:
            # Caminho do banco atual
            db_path = settings.DATABASES['default']['NAME']
            
            # Criar backup do banco atual antes de restaurar
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_atual = os.path.join(backups_dir, f'pre_restore_{timestamp}.db')
            shutil.copy2(db_path, backup_atual)
            
            # Restaurar o backup selecionado
            shutil.copy2(backup_path, db_path)
            
            messages.success(request, 'Backup restaurado com sucesso! Recarregue a página.')
            
        except Exception as e:
            messages.error(request, f'Erro ao restaurar backup: {str(e)}')
    
    return redirect('core:backup_restore')


@login_required
def backup_download(request, backup_name):
    """Baixar arquivo de backup"""
    from django.conf import settings
    import os
    from django.http import FileResponse

    backups_dir = os.path.realpath(os.path.join(settings.BASE_DIR, 'backups'))
    backup_path = os.path.realpath(os.path.join(backups_dir, backup_name))

    if not backup_path.startswith(backups_dir + os.sep) or not backup_name.endswith('.db'):
        messages.error(request, 'Acesso negado!')
        return redirect('core:backup_restore')

    if os.path.exists(backup_path):
        response = FileResponse(open(backup_path, 'rb'), as_attachment=True, filename=backup_name)
        return response
    else:
        messages.error(request, 'Arquivo de backup não encontrado!')
        return redirect('core:backup_restore')


@login_required
def backup_deletar(request, backup_name):
    """Deletar arquivo de backup"""
    from django.conf import settings
    import os

    backups_dir = os.path.realpath(os.path.join(settings.BASE_DIR, 'backups'))
    backup_path = os.path.realpath(os.path.join(backups_dir, backup_name))

    if not backup_path.startswith(backups_dir + os.sep) or not backup_name.endswith('.db'):
        messages.error(request, 'Acesso negado!')
        return redirect('core:backup_restore')

    try:
        if os.path.exists(backup_path):
            os.remove(backup_path)
            messages.success(request, 'Backup deletado com sucesso!')
        else:
            messages.error(request, 'Arquivo de backup não encontrado!')

    except Exception as e:
        messages.error(request, f'Erro ao deletar backup: {str(e)}')

    return redirect('core:backup_restore')


def verificar_licenca(request):
    """View para verificar status da licença e permitir ativação"""
    from .models import Licenca
    from django.utils import timezone
    
    licenca = Licenca.licenca_valida()
    
    if licenca:
        # Licença válida - redirecionar para dashboard
        return redirect('core:dashboard')
    
    # Licença expirada ou não existe
    if request.method == 'POST':
        chave = request.POST.get('chave', '').strip().upper()
        
        sucesso, resultado = Licenca.ativar_licenca(chave)
        
        if sucesso:
            messages.success(request, f'Licença ativada com sucesso! Válida até {resultado.data_expiracao.strftime("%d/%m/%Y")}')
            return redirect('core:dashboard')
        else:
            messages.error(request, resultado)
    
    # Mostrar tela de licença expirada
    return render(request, 'core/licenca_expirada.html', {
        'licenca_anterior': Licenca.objects.filter(ativa=False).latest('data_ativacao') if Licenca.objects.filter(ativa=False).exists() else None
    })


def gerar_chave_licenca(request):
    """Gerar nova chave de licença (requer autenticação admin)"""
    import hashlib
    import secrets
    import string
    from datetime import datetime
    
    if not request.user.is_superuser:
        messages.error(request, 'Acesso negado!')
        return redirect('core:dashboard')
    
    # Gerar chave: 4 blocos de 4 caracteres + checksum
    def gerar_bloco():
        chars = string.ascii_uppercase + string.digits
        return ''.join(secrets.choice(chars) for _ in range(4))
    
    blocos = [gerar_bloco() for _ in range(4)]
    chave_base = '-'.join(blocos)
    
    # Calcular checksum
    hash_value = hashlib.md5(chave_base.encode()).hexdigest()[:4].upper()
    chave_completa = f"{chave_base}-{hash_value}"
    
    data_geracao = datetime.now().strftime('%d/%m/%Y')
    
    context = {
        'chave': chave_completa,
        'data_geracao': data_geracao,
        'validade': 90  # dias
    }
    
    return render(request, 'core/gerar_chave_licenca.html', context)


@login_required
def termo_compra(request, compra_id):
    """Gerar termo de compra de produtos usados"""
    from .models import Compra
    
    compra = Compra.objects.get(id=compra_id)
    emitente = Emitente.objects.first()
    
    context = {
        'compra': compra,
        'emitente': emitente,
    }
    
    return render(request, 'core/termo_compra.html', context)


@login_required
def termo_venda(request, venda_id):
    """Gerar termo de venda para impressão"""
    venda = get_object_or_404(Venda, id=venda_id)
    emitente = Emitente.objects.first()

    # Mêses de garantia contratual: pode ser passado via GET (?meses=6) ou padrão 3
    try:
        meses_garantia = int(request.GET.get('meses', 3))
    except (ValueError, TypeError):
        meses_garantia = 3

    context = {
        'venda': venda,
        'emitente': emitente,
        'itens': venda.items.all(),
        'tradeins': venda.tradeins.all(),
        'meses_garantia': meses_garantia,
    }
    return render(request, 'core/termo_venda.html', context)


@login_required
def termo_garantia_conserto(request, os_id):
    """Gerar termo de garantia de conserto (OS finalizada)"""
    from django.utils import timezone
    os_obj = get_object_or_404(OrdemServico, id=os_id)
    emitente = Emitente.objects.first()

    # Garante que a contagem começa a partir do fechamento/entrega ao cliente
    # Prioridade: data_entrega > data_conclusao > hoje
    from django.utils import timezone
    data_base = os_obj.data_entrega or os_obj.data_conclusao or timezone.now().date()
    data_fim_garantia = data_base + timedelta(days=os_obj.garantia_dias)

    context = {
        'os': os_obj,
        'emitente': emitente,
        'data_inicio_garantia': data_base,
        'data_fim_garantia': data_fim_garantia,
        'prazo_atendimento': 3,
        'now': timezone.now(),
    }
    return render(request, 'core/termo_garantia_conserto.html', context)


# ============ ORDENS DE SERVIÇO ============

@login_required
def lista_os(request):
    """Listar todas as Ordens de Serviço"""
    ordens = OrdemServico.objects.select_related('cliente').order_by('-data_entrada')

    # Filtros
    status = request.GET.get('status', '')
    tipo = request.GET.get('tipo', '')
    busca = request.GET.get('busca', '')

    if status:
        ordens = ordens.filter(status=status)
    if tipo:
        ordens = ordens.filter(tipo_servico=tipo)
    if busca:
        ordens = ordens.filter(
            Q(numero__icontains=busca) |
            Q(cliente__nome__icontains=busca) |
            Q(aparelho__icontains=busca) |
            Q(imei__icontains=busca)
        )

    paginator = Paginator(ordens, 20)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page,
        'ordens': page,  # compatibilidade com template
        'status_choices': OrdemServico.STATUS_CHOICES,
        'tipo_choices': OrdemServico.TIPO_SERVICO_CHOICES,
        'status_filtro': status,
        'tipo_filtro': tipo,
        'busca': busca,
        'total': ordens.count(),
        'aguardando': OrdemServico.objects.filter(status='aguardando').count(),
        'em_andamento': OrdemServico.objects.filter(status='em_andamento').count(),
        'concluido': OrdemServico.objects.filter(status='concluido').count(),
    }
    return render(request, 'core/lista_os.html', context)


@login_required
def nova_os(request):
    """Criar nova Ordem de Serviço"""
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        if not cliente_id:
            messages.error(request, 'Selecione um cliente!')
            return render(request, 'core/nova_os.html', {
                'clientes': Cliente.objects.order_by('nome'),
                'tipo_choices': OrdemServico.TIPO_SERVICO_CHOICES,
                'post': request.POST,
            })
        try:
            os_obj = OrdemServico.objects.create(
                cliente_id=cliente_id,
                aparelho=request.POST.get('aparelho', ''),
                imei=request.POST.get('imei', ''),
                imei2=request.POST.get('imei2', ''),
                acessorios=request.POST.get('acessorios', ''),
                tipo_servico=request.POST.get('tipo_servico', 'outro'),
                descricao_problema=request.POST.get('descricao_problema', ''),
                observacoes_tecnico=request.POST.get('observacoes_tecnico', ''),
                valor_orcamento=Decimal(str(request.POST.get('valor_orcamento', '0') or '0')),
                data_previsao=request.POST.get('data_previsao') or None,
                tecnico=request.POST.get('tecnico', request.user.username),
                garantia_dias=int(request.POST.get('garantia_dias', 90)),
            )
            os_obj.refresh_from_db()
            messages.success(request, f'OS #{os_obj.numero} criada com sucesso!')

            # ── Enviar WhatsApp automaticamente ──────────────────
            telefone = (os_obj.cliente.telefone or '').strip() if os_obj.cliente else ''
            if telefone:
                previsao = (
                    os_obj.data_previsao.strftime('%d/%m/%Y')
                    if os_obj.data_previsao else 'a confirmar'
                )
                orcamento = (
                    f'R$ {os_obj.valor_orcamento:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                    if os_obj.valor_orcamento else 'a definir'
                )
                primeiro_nome = os_obj.cliente.nome.split()[0]
                msg = (
                    f'Ola, {primeiro_nome}!\n\n'
                    f'Sua Ordem de Servico foi aberta com sucesso!\n\n'
                    f'OS #{os_obj.numero}\n'
                    f'Aparelho: {os_obj.aparelho}\n'
                    f'Servico: {os_obj.get_tipo_servico_display()}\n'
                    f'Problema: {os_obj.descricao_problema[:120]}\n'
                    f'Orcamento: {orcamento}\n'
                    f'Previsao: {previsao}\n\n'
                    f'Aguarde nosso contato. Obrigado!'
                )
                ok, erro = _wa_send(telefone, msg)
                if ok:
                    messages.info(request, f'WhatsApp enviado para {os_obj.cliente.nome}!')
                else:
                    messages.warning(request, f'OS criada, mas WhatsApp nao enviado: {erro}')
            # ─────────────────────────────────────────────────────

            return redirect('core:detalhe_os', os_id=os_obj.id)
        except Exception as e:
            messages.error(request, f'Erro ao criar OS: {str(e)}')

    return render(request, 'core/nova_os.html', {
        'clientes': Cliente.objects.order_by('nome'),
        'tipo_choices': OrdemServico.TIPO_SERVICO_CHOICES,
    })


@login_required
def detalhe_os(request, os_id):
    """Ver detalhes e atualizar status da OS"""
    os_obj = get_object_or_404(OrdemServico, id=os_id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'atualizar_status':
            novo_status = request.POST.get('status')
            obs_tecnico = request.POST.get('observacoes_tecnico', os_obj.observacoes_tecnico)
            valor_final = request.POST.get('valor_final', '')
            forma_pagamento = request.POST.get('forma_pagamento', '')
            obs_historico = request.POST.get('obs_historico', '').strip()

            try:
                status_anterior = os_obj.status
                os_obj.status = novo_status
                os_obj.observacoes_tecnico = obs_tecnico
                if valor_final:
                    os_obj.valor_final = Decimal(str(valor_final))
                if forma_pagamento:
                    os_obj.forma_pagamento = forma_pagamento
                if novo_status == 'concluido' and not os_obj.data_conclusao:
                    from django.utils import timezone
                    os_obj.data_conclusao = timezone.now().date()
                if novo_status == 'entregue' and not os_obj.data_entrega:
                    from django.utils import timezone
                    os_obj.data_entrega = timezone.now().date()
                os_obj.save()

                # Grava histórico apenas se o status mudou
                from .models import HistoricoStatusOS
                if status_anterior != novo_status:
                    HistoricoStatusOS.objects.create(
                        os=os_obj,
                        status_anterior=status_anterior,
                        status_novo=novo_status,
                        usuario=request.user,
                        observacao=obs_historico,
                    )

                messages.success(request, f'OS #{os_obj.numero} atualizada para "{os_obj.get_status_display()}"')
            except Exception as e:
                messages.error(request, f'Erro ao atualizar OS: {str(e)}')
            return redirect('core:detalhe_os', os_id=os_id)

        if action == 'editar':
            os_obj.aparelho = request.POST.get('aparelho', os_obj.aparelho)
            os_obj.imei = request.POST.get('imei', os_obj.imei)
            os_obj.imei2 = request.POST.get('imei2', os_obj.imei2)
            os_obj.acessorios = request.POST.get('acessorios', os_obj.acessorios)
            os_obj.tipo_servico = request.POST.get('tipo_servico', os_obj.tipo_servico)
            os_obj.descricao_problema = request.POST.get('descricao_problema', os_obj.descricao_problema)
            os_obj.valor_orcamento = Decimal(str(request.POST.get('valor_orcamento', os_obj.valor_orcamento) or '0'))
            os_obj.data_previsao = request.POST.get('data_previsao') or os_obj.data_previsao
            os_obj.tecnico = request.POST.get('tecnico', os_obj.tecnico)
            os_obj.garantia_dias = int(request.POST.get('garantia_dias', os_obj.garantia_dias))
            os_obj.save()
            messages.success(request, 'OS atualizada com sucesso!')
            return redirect('core:detalhe_os', os_id=os_id)

    from .models import HistoricoStatusOS
    context = {
        'os': os_obj,
        'status_choices': OrdemServico.STATUS_CHOICES,
        'tipo_choices': OrdemServico.TIPO_SERVICO_CHOICES,
        'forma_pagamento_choices': [
            ('dinheiro', 'Dinheiro'), ('debito', 'Débito'), ('credito', 'Crédito'),
            ('pix', 'PIX'), ('parcelado', 'Parcelado'),
        ],
        'historico': HistoricoStatusOS.objects.filter(os=os_obj).select_related('usuario').order_by('data_hora'),
    }
    return render(request, 'core/detalhe_os.html', context)


@login_required
def imprimir_os(request, os_id):
    """Imprimir / visualizar comprovante da OS"""
    os_obj = get_object_or_404(OrdemServico, id=os_id)
    emitente = Emitente.objects.first()
    return render(request, 'core/imprimir_os.html', {'os': os_obj, 'emitente': emitente})


# ============ HISTÓRICO DO CLIENTE ============

@login_required
def detalhe_cliente(request, cliente_id):
    """Histórico completo do cliente: compras, garantias ativas e OSs abertas"""
    cliente = get_object_or_404(Cliente, id=cliente_id)

    vendas = (
        Venda.objects
        .filter(cliente=cliente, status='finalizada')
        .prefetch_related('items__produto')
        .order_by('-data_venda')
    )
    garantias_ativas = (
        Garantia.objects
        .filter(venda__cliente=cliente, data_fim__gte=datetime.now().date())
        .select_related('produto', 'venda')
        .order_by('data_fim')
    )
    ordens = (
        OrdemServico.objects
        .filter(cliente=cliente)
        .order_by('-data_entrada')
    )

    total_gasto = vendas.aggregate(Sum('total'))['total__sum'] or Decimal('0')
    total_vendas = vendas.count()
    os_abertas = ordens.exclude(status='concluido').count()

    context = {
        'cliente': cliente,
        'vendas': vendas,
        'garantias': garantias_ativas,
        'ordens_servico': ordens,
        'total_gasto': total_gasto,
        'total_vendas': total_vendas,
        'os_abertas': os_abertas,
    }
    return render(request, 'core/detalhe_cliente.html', context)


@login_required
def editar_cliente(request, cliente_id):
    """Editar cliente com Django Form"""
    cliente = get_object_or_404(Cliente, id=cliente_id)
    form = ClienteForm(request.POST or None, instance=cliente)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, 'Cliente atualizado com sucesso!')
            return redirect('core:detalhe_cliente', cliente_id=cliente.id)
        except Exception as e:
            from django.db import IntegrityError
            if isinstance(e, IntegrityError) and 'unique' in str(e).lower():
                messages.error(request, 'CPF ou e-mail já cadastrado para outro cliente.')
            else:
                messages.error(request, f'Erro ao atualizar cliente: {str(e)}')
    return render(request, 'core/cadastro_cliente.html', {'form': form, 'cliente': cliente})


@login_required
def editar_produto(request, produto_id):
    """Editar produto com Django Form"""
    produto = get_object_or_404(Produto, id=produto_id)
    form = ProdutoForm(request.POST or None, request.FILES or None, instance=produto)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, 'Produto atualizado com sucesso!')
            return redirect('core:listar_produtos')
        except Exception as e:
            messages.error(request, f'Erro ao atualizar produto: {str(e)}')
    return render(request, 'core/cadastro_produto.html', {'form': form, 'produto': produto})


# ============ API: IMEIs por produto (Ajax) ============

@login_required
def api_imeis_produto(request, produto_id):
    """Retorna em JSON os SerialProduto em estoque para o produto dado"""
    from django.http import JsonResponse
    seriais = list(
        SerialProduto.objects.filter(produto_id=produto_id, status='em_estoque')
        .values('id', 'serial', 'imei', 'imei2', 'observacoes')
    )
    return JsonResponse({'seriais': seriais})


# ============ BUSCA POR IMEI ============

@login_required
def busca_imei(request):
    """Busca rápida por IMEI 1 ou IMEI 2: encontra garantias e OSs associadas"""
    imei1_q = request.GET.get('imei1', '').strip()
    imei2_q = request.GET.get('imei2', '').strip()
    garantia = None
    os_resultado = None
    erro = None
    imei_buscado = imei1_q or imei2_q

    if imei1_q or imei2_q:
        if imei1_q and (not imei1_q.isdigit() or len(imei1_q) != 15):
            erro = 'IMEI 1 deve conter exatamente 15 dígitos numéricos.'
        elif imei2_q and (not imei2_q.isdigit() or len(imei2_q) != 15):
            erro = 'IMEI 2 deve conter exatamente 15 dígitos numéricos.'
        else:
            # Monta filtro de acordo com qual campo foi preenchido
            if imei1_q and imei2_q:
                filtro_g = Q(imei=imei1_q) | Q(imei2=imei2_q)
                filtro_o = Q(imei=imei1_q) | Q(imei2=imei2_q)
            elif imei1_q:
                filtro_g = Q(imei=imei1_q)
                filtro_o = Q(imei=imei1_q)
            else:
                filtro_g = Q(imei2=imei2_q)
                filtro_o = Q(imei2=imei2_q)

            garantia = (
                Garantia.objects
                .filter(filtro_g)
                .select_related('venda__cliente', 'produto')
                .first()
            )
            os_resultado = (
                OrdemServico.objects
                .filter(filtro_o)
                .select_related('cliente')
                .order_by('-data_entrada')
                .first()
            )
            if not garantia and not os_resultado:
                campo = 'IMEI 1' if imei1_q and not imei2_q else ('IMEI 2' if imei2_q and not imei1_q else 'IMEIs')
                erro = f'Nenhum registro encontrado para {campo}: {imei_buscado}.'

    return render(request, 'core/busca_imei.html', {
        'imei1': imei1_q,
        'imei2': imei2_q,
        'garantia': garantia,
        'ordem_servico': os_resultado,
        'erro': erro,
        'hoje': datetime.now().date(),
    })


# ============ CANCELAMENTO DE VENDA ============

@login_required
def cancelar_venda(request, venda_id):
    """Cancela a venda e faz estorno de estoque para vendas finalizadas"""
    venda = get_object_or_404(Venda, id=venda_id)

    if venda.status == 'cancelada':
        messages.warning(request, 'Esta venda já foi cancelada.')
        return redirect('core:venda_visualizar', venda_id=venda_id)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Estornar estoque apenas se venda estava finalizada
                if venda.status == 'finalizada':
                    for item in venda.items.select_related('produto').all():
                        item.produto.estoque += item.quantidade
                        item.produto.save(update_fields=['estoque'])

                venda.status = 'cancelada'
                venda.save(update_fields=['status'])

            messages.success(request, f'Venda #{venda.id} cancelada. Estoque restaurado.')
        except Exception as e:
            messages.error(request, f'Erro ao cancelar venda: {str(e)}')

    return redirect('core:venda_visualizar', venda_id=venda_id)


# ============================================================
#  CÓDIGO DE BARRAS EAN-13
# ============================================================

@login_required
def barcode_produto(request, produto_id):
    """Gera imagem PNG do código de barras do produto (EAN-13)"""
    import barcode
    from barcode.writer import ImageWriter
    import io

    produto = get_object_or_404(Produto, id=produto_id)
    codigo = produto.codigo_barras.strip() if produto.codigo_barras else None
    if not codigo:
        return HttpResponse('Produto sem código de barras cadastrado.', status=400)

    try:
        writer = ImageWriter()
        if len(codigo) == 13 and codigo.isdigit():
            cls = barcode.get_barcode_class('ean13')
            bc = cls(codigo, writer=writer)
        else:
            cls = barcode.get_barcode_class('code128')
            bc = cls(codigo, writer=writer)
        buf = io.BytesIO()
        bc.write(buf, options={'write_text': True, 'quiet_zone': 6})
        buf.seek(0)
        return HttpResponse(buf.read(), content_type='image/png')
    except Exception as e:
        return HttpResponse(f'Erro ao gerar barcode: {e}', status=500)


# ============================================================
#  RELATÓRIO DE LUCRO POR PERÍODO
# ============================================================

@login_required
def relatorio_lucro(request):
    """Relatório de custo × receita × lucro bruto por período"""
    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')

    hoje = datetime.now().date()
    try:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else hoje - timedelta(days=30)
    except ValueError:
        data_inicio = hoje - timedelta(days=30)
    try:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else hoje
    except ValueError:
        data_fim = hoje

    vendas_periodo = Venda.objects.filter(
        status='finalizada',
        data_venda__date__range=(data_inicio, data_fim)
    ).prefetch_related('items__produto').order_by('-data_venda')

    receita_total = Decimal('0')
    custo_total = Decimal('0')
    linhas = []

    for venda in vendas_periodo:
        receita_venda = venda.total
        custo_venda = sum(
            item.produto.preco_custo * item.quantidade
            for item in venda.items.all()
        )
        lucro_venda = receita_venda - custo_venda
        receita_total += receita_venda
        custo_total += custo_venda
        linhas.append({
            'venda': venda,
            'receita': receita_venda,
            'custo': custo_venda,
            'lucro': lucro_venda,
            'margem': float(lucro_venda / receita_venda * 100) if receita_venda else 0,
        })

    lucro_total = receita_total - custo_total
    margem_media = float(lucro_total / receita_total * 100) if receita_total else 0

    return render(request, 'core/relatorio_lucro.html', {
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'linhas': linhas,
        'receita_total': receita_total,
        'custo_total': custo_total,
        'lucro_total': lucro_total,
        'margem_media': margem_media,
        'emitente': Emitente.get_ativo(),
    })


# ============================================================
#  EXPORTAÇÃO PARA EXCEL (openpyxl)
# ============================================================

@login_required
def exportar_excel(request, tipo):
    """Exporta dados para Excel (.xlsx)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active

    def set_header(ws, row_num=1):
        for cell in ws[row_num]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E79')
            cell.alignment = Alignment(horizontal='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    if tipo == 'vendas':
        ws.title = 'Vendas'
        ws.append(['ID', 'Data', 'Cliente', 'CPF', 'Vendedor', 'Status', 'Forma Pag.', 'Total (R$)'])
        set_header(ws)
        for v in Venda.objects.select_related('cliente').order_by('-data_venda'):
            ws.append([v.id, v.data_venda.strftime('%d/%m/%Y %H:%M'), v.cliente.nome, v.cliente.cpf,
                       v.vendedor, v.get_status_display(), v.get_forma_pagamento_display(), float(v.total)])
        nome_arquivo = 'vendas'

    elif tipo == 'produtos':
        ws.title = 'Produtos'
        ws.append(['ID', 'Nome', 'Tipo', 'Marca', 'Modelo', 'Cód. Barras', 'Custo (R$)', 'Venda (R$)', 'Estoque', 'Est. Mín.'])
        set_header(ws)
        for p in Produto.objects.order_by('nome'):
            ws.append([p.id, p.nome, p.get_tipo_display(), p.marca, p.modelo, p.codigo_barras,
                       float(p.preco_custo), float(p.preco_venda), p.estoque, p.estoque_minimo])
        nome_arquivo = 'produtos'

    elif tipo == 'clientes':
        ws.title = 'Clientes'
        ws.append(['ID', 'Nome', 'CPF', 'Telefone', 'E-mail', 'Endereço'])
        set_header(ws)
        for c in Cliente.objects.order_by('nome'):
            ws.append([c.id, c.nome, c.cpf, c.telefone, c.email, c.endereco])
        nome_arquivo = 'clientes'

    elif tipo == 'os':
        ws.title = 'OrdemServico'
        ws.append(['Nº OS', 'Data', 'Cliente', 'Aparelho', 'IMEI', 'Serviço', 'Técnico', 'Status', 'Orçamento (R$)', 'Final (R$)'])
        set_header(ws)
        for os in OrdemServico.objects.select_related('cliente').order_by('-data_entrada'):
            ws.append([os.numero, os.data_entrada.strftime('%d/%m/%Y'), os.cliente.nome,
                       os.aparelho, os.imei, os.get_tipo_servico_display(), os.tecnico,
                       os.get_status_display(), float(os.valor_orcamento), float(os.valor_final)])
        nome_arquivo = 'ordens_servico'

    else:
        return HttpResponse('Tipo inválido', status=400)

    auto_width(ws)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{nome_arquivo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


# ============================================================
#  WHATSAPP VIA BAILEYS
# ============================================================

def _wa_send(telefone, mensagem):
    """
    Envia mensagem WhatsApp via microservico Baileys.
    Retorna (True, None) em sucesso ou (False, motivo) em falha.
    Nao lanca excecoes — apenas retorna o resultado.
    """
    import urllib.request as _req
    import json as _j
    from django.conf import settings
    WA_URL = getattr(settings, 'WHATSAPP_SERVICE_URL', 'http://localhost:3000')
    numero = ''.join(filter(str.isdigit, telefone))
    if not numero.startswith('55'):
        numero = '55' + numero
    try:
        payload = _j.dumps({'phone': numero, 'message': mensagem}).encode('utf-8')
        req = _req.Request(f'{WA_URL}/send', data=payload,
                           headers={'Content-Type': 'application/json'}, method='POST')
        with _req.urlopen(req, timeout=8) as resp:
            data = _j.loads(resp.read())
        ok = data.get('success', False)
        return ok, data.get('error') if not ok else None
    except Exception as e:
        return False, str(e)


@login_required
def whatsapp_enviar(request):
    """Envia mensagem WhatsApp via microservico Node.js/Baileys"""
    import urllib.request as urlreq
    import json as _json
    from django.conf import settings
    from django.http import JsonResponse

    if request.method != 'POST':
        return HttpResponse(status=405)

    WA_URL = getattr(settings, 'WHATSAPP_SERVICE_URL', 'http://localhost:3000')
    telefone = request.POST.get('telefone', '').strip()
    mensagem = request.POST.get('mensagem', '').strip()
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    next_url = request.POST.get('next', request.META.get('HTTP_REFERER', '/'))

    if not telefone or not mensagem:
        if is_ajax:
            return JsonResponse({'ok': False, 'error': 'Telefone e mensagem são obrigatórios.'})
        messages.error(request, 'Telefone e mensagem são obrigatórios.')
        return redirect(next_url)

    numero = ''.join(filter(str.isdigit, telefone))
    if not numero.startswith('55'):
        numero = '55' + numero

    try:
        payload = _json.dumps({'phone': numero, 'message': mensagem}).encode('utf-8')
        req = urlreq.Request(f'{WA_URL}/send', data=payload,
                             headers={'Content-Type': 'application/json'}, method='POST')
        with urlreq.urlopen(req, timeout=8) as resp:
            data = _json.loads(resp.read())
        ok = data.get('success', False)
        if is_ajax:
            return JsonResponse({'ok': ok, 'error': data.get('error', '')})
        if ok:
            messages.success(request, f'WhatsApp enviado para {telefone}!')
        else:
            messages.error(request, f"Falha no envio: {data.get('error', 'erro desconhecido')}")
    except Exception as e:
        if is_ajax:
            return JsonResponse({'ok': False, 'error': str(e)})
        messages.error(request, f'Serviço WhatsApp indisponível: {e}')

    return redirect(next_url)


@login_required
def whatsapp_status(request):
    """Verifica status da conexão WhatsApp"""
    import urllib.request as urlreq
    import json as _json
    from django.conf import settings
    from django.http import JsonResponse

    WA_URL = getattr(settings, 'WHATSAPP_SERVICE_URL', 'http://localhost:3000')
    try:
        req = urlreq.Request(f'{WA_URL}/status', method='GET')
        with urlreq.urlopen(req, timeout=4) as resp:
            data = _json.loads(resp.read())
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'connected': False, 'error': str(e)})


@login_required
def whatsapp_painel(request):
    """Painel de conexão WhatsApp com QR Code"""
    return render(request, 'core/whatsapp_painel.html')


@login_required
def whatsapp_qr(request):
    """Retorna o QR Code atual do serviço WhatsApp"""
    import urllib.request as urlreq
    import json as _json
    from django.conf import settings
    from django.http import JsonResponse

    WA_URL = getattr(settings, 'WHATSAPP_SERVICE_URL', 'http://localhost:3000')
    try:
        req = urlreq.Request(f'{WA_URL}/qr', method='GET')
        with urlreq.urlopen(req, timeout=4) as resp:
            data = _json.loads(resp.read())
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=503)


@login_required
def whatsapp_conectar(request):
    """Inicia/reinicia conexão WhatsApp"""
    import urllib.request as urlreq
    import json as _json
    from django.conf import settings
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)

    WA_URL = getattr(settings, 'WHATSAPP_SERVICE_URL', 'http://localhost:3000')
    try:
        payload = b'{}'
        req = urlreq.Request(f'{WA_URL}/connect', data=payload,
                             headers={'Content-Type': 'application/json'}, method='POST')
        with urlreq.urlopen(req, timeout=6) as resp:
            data = _json.loads(resp.read())
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=503)


@login_required
def whatsapp_desconectar(request):
    """Desconecta a sessão WhatsApp"""
    import urllib.request as urlreq
    import json as _json
    from django.conf import settings
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)

    WA_URL = getattr(settings, 'WHATSAPP_SERVICE_URL', 'http://localhost:3000')
    try:
        req = urlreq.Request(f'{WA_URL}/disconnect', method='DELETE')
        with urlreq.urlopen(req, timeout=6) as resp:
            data = _json.loads(resp.read())
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=503)


# ============================================================
#  CONTROLE DE SERIAL / IMEI POR PRODUTO
# ============================================================

@login_required
def serial_listar(request, produto_id):
    """Lista e cadastra seriais/IMEIs de um produto"""
    produto = get_object_or_404(Produto, id=produto_id)
    seriais = produto.seriais.all()
    form = SerialForm()
    if request.method == 'POST':
        form = SerialForm(request.POST)
        if form.is_valid():
            serial = form.save(commit=False)
            serial.produto = produto
            serial.save()
            messages.success(request, f'Serial {serial.serial} cadastrado!')
            return redirect('core:serial_listar', produto_id=produto.id)
    return render(request, 'core/lista_seriais.html', {
        'produto': produto,
        'seriais': seriais,
        'form': form,
    })


@login_required
def serial_deletar(request, serial_id):
    """Remove um serial cadastrado"""
    from django.http import JsonResponse
    serial = get_object_or_404(SerialProduto, id=serial_id)
    produto_id = serial.produto_id
    if request.method == 'POST':
        serial.delete()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True})
        messages.success(request, 'Serial removido.')
    return redirect('core:serial_listar', produto_id=produto_id)


# ============================================================
#  CRUD DE FORNECEDORES
# ============================================================

@login_required
@verificar_licenca_view
@login_required
@verificar_licenca_view
def listar_fornecedores(request):
    """Listar fornecedores com busca"""
    qs = Fornecedor.objects.all().order_by('nome')
    busca = request.GET.get('busca', '').strip()
    if busca:
        qs = qs.filter(Q(nome__icontains=busca) | Q(cnpj__icontains=busca) | Q(telefone__icontains=busca) | Q(contato__icontains=busca))
    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/lista_fornecedores.html', {'page_obj': page, 'busca': busca})


@login_required
@verificar_licenca_view
def detalhe_fornecedor(request, fornecedor_id):
    """Exibe detalhes do fornecedor com histórico de compras e produtos"""
    fornecedor = get_object_or_404(Fornecedor, id=fornecedor_id)
    from .models import Compra
    compras = Compra.objects.filter(fornecedor=fornecedor).order_by('-data_compra').select_related()[:30]
    produtos = Produto.objects.filter(fornecedor=fornecedor, ativo=True).order_by('nome')
    total_compras = Compra.objects.filter(fornecedor=fornecedor).count()
    return render(request, 'core/detalhe_fornecedor.html', {
        'fornecedor': fornecedor,
        'compras': compras,
        'produtos': produtos,
        'total_compras': total_compras,
    })


@login_required
@verificar_licenca_view
def cadastrar_fornecedor(request):
    """Cadastrar novo fornecedor"""
    form = FornecedorForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        try:
            f = form.save()
            messages.success(request, f'Fornecedor "{f.nome}" cadastrado com sucesso!')
            return redirect('core:listar_fornecedores')
        except Exception as e:
            from django.db import IntegrityError
            if isinstance(e, IntegrityError) and 'unique' in str(e).lower():
                messages.error(request, 'CNPJ já cadastrado para outro fornecedor.')
            else:
                messages.error(request, f'Erro ao salvar fornecedor: {str(e)}')
    return render(request, 'core/cadastro_fornecedor.html', {'form': form, 'titulo': 'Novo Fornecedor'})


@login_required
@verificar_licenca_view
def editar_fornecedor(request, fornecedor_id):
    """Editar fornecedor existente"""
    fornecedor = get_object_or_404(Fornecedor, id=fornecedor_id)
    form = FornecedorForm(request.POST or None, instance=fornecedor)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, f'Fornecedor "{fornecedor.nome}" atualizado!')
            return redirect('core:listar_fornecedores')
        except Exception as e:
            messages.error(request, f'Erro ao atualizar fornecedor: {str(e)}')
    return render(request, 'core/cadastro_fornecedor.html', {
        'form': form, 'titulo': f'Editar: {fornecedor.nome}', 'fornecedor': fornecedor
    })


@login_required
@verificar_licenca_view
def deletar_fornecedor(request, fornecedor_id):
    """Deletar fornecedor (apenas POST)"""
    fornecedor = get_object_or_404(Fornecedor, id=fornecedor_id)
    if request.method == 'POST':
        try:
            nome = fornecedor.nome
            fornecedor.delete()
            messages.success(request, f'Fornecedor "{nome}" removido.')
        except Exception:
            messages.error(request, 'Não é possível remover: existem produtos ou compras vinculadas.')
    return redirect('core:listar_fornecedores')


# ============================================================
#  RELATÓRIO DE ORDENS DE SERVIÇO
# ============================================================

@login_required
@verificar_licenca_view
def relatorio_os(request):
    """Relatório de OS por período, técnico e tipo"""
    hoje = datetime.now().date()
    data_ini = request.GET.get('data_ini', (hoje - timedelta(days=30)).isoformat())
    data_fim = request.GET.get('data_fim', hoje.isoformat())
    tecnico = request.GET.get('tecnico', '').strip()
    status = request.GET.get('status', '')
    tipo = request.GET.get('tipo', '')

    qs = OrdemServico.objects.select_related('cliente').filter(
        data_entrada__date__gte=data_ini,
        data_entrada__date__lte=data_fim,
    )
    if tecnico:
        qs = qs.filter(tecnico__icontains=tecnico)
    if status:
        qs = qs.filter(status=status)
    if tipo:
        qs = qs.filter(tipo_servico=tipo)

    # Totalizadores
    total_os = qs.count()
    total_orcamento = qs.aggregate(s=Sum('valor_orcamento'))['s'] or Decimal('0')
    total_faturado = qs.filter(
        status__in=['concluido', 'entregue']
    ).aggregate(s=Sum('valor_final'))['s'] or Decimal('0')

    # Agrupamentos
    por_status = (
        qs.values('status')
        .annotate(qtd=Count('id'), faturado=Sum('valor_final'))
        .order_by('-qtd')
    )
    por_tipo = (
        qs.values('tipo_servico')
        .annotate(qtd=Count('id'))
        .order_by('-qtd')
    )
    por_tecnico = (
        qs.values('tecnico')
        .annotate(qtd=Count('id'), faturado=Sum('valor_final'))
        .order_by('-qtd')[:10]
    )

    # Traduzir status nos agrupamentos
    status_display = dict(OrdemServico.STATUS_CHOICES)
    tipo_display = dict(OrdemServico.TIPO_SERVICO_CHOICES)
    for item in por_status:
        item['status_label'] = status_display.get(item['status'], item['status'])
    for item in por_tipo:
        item['tipo_label'] = tipo_display.get(item['tipo_servico'], item['tipo_servico'])

    paginator = Paginator(qs.order_by('-data_entrada'), 25)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page,
        'data_ini': data_ini,
        'data_fim': data_fim,
        'tecnico': tecnico,
        'status_filtro': status,
        'tipo_filtro': tipo,
        'total_os': total_os,
        'total_orcamento': total_orcamento,
        'total_faturado': total_faturado,
        'por_status': por_status,
        'por_tipo': por_tipo,
        'por_tecnico': por_tecnico,
        'status_choices': OrdemServico.STATUS_CHOICES,
        'tipo_choices': OrdemServico.TIPO_SERVICO_CHOICES,
        'emitente': Emitente.get_ativo(),
    }
    return render(request, 'core/relatorio_os.html', context)


# ============================================================
#  OS → VENDA (Converter OS em cobrança de venda)
# ============================================================

@login_required
@verificar_licenca_view
def os_gerar_venda(request, os_id):
    """Converte uma OS concluída em uma Venda registrada no sistema"""
    os_obj = get_object_or_404(OrdemServico, id=os_id)

    if os_obj.status not in ('concluido', 'entregue'):
        messages.error(request, 'A OS precisa estar Concluída para gerar uma venda.')
        return redirect('core:detalhe_os', os_id=os_id)

    if not os_obj.valor_final or os_obj.valor_final <= 0:
        messages.error(request, 'Preencha o Valor Final da OS antes de gerar a venda.')
        return redirect('core:detalhe_os', os_id=os_id)

    # Verificar se já existe venda gerada desta OS (usando observacoes como link)
    venda_existente = Venda.objects.filter(
        observacoes__icontains=f'OS #{os_obj.numero}'
    ).first()
    if venda_existente:
        messages.warning(request, f'Já existe a Venda #{venda_existente.id} para esta OS.')
        return redirect('core:venda_visualizar', venda_id=venda_existente.id)

    if request.method == 'POST':
        forma_pagamento = request.POST.get('forma_pagamento', 'dinheiro')
        parcelas = int(request.POST.get('parcelas', 1) or 1)
        vendedor = request.POST.get('vendedor', request.user.username)

        with transaction.atomic():
            venda = Venda.objects.create(
                cliente=os_obj.cliente,
                vendedor=vendedor,
                status='finalizada',
                total=os_obj.valor_final,
                forma_pagamento=forma_pagamento,
                parcelas=parcelas,
                observacoes=f'Gerado automaticamente da OS #{os_obj.numero}',
            )
            # Se houver produto principal na OS (via aparelho), usamos produto genérico de serviço
            # Para a OS usamos um produto "Serviço" — buscamos o primeiro produto de serviço
            # ou criamos uma linha com preço manual sem produto estoque
            # Solução: registrar como produto de mão-de-obra se existir, senão sem item
            produto_servico = Produto.objects.filter(
                nome__icontains='serviço', ativo=True
            ).first() or Produto.objects.filter(
                nome__icontains='mão de obra', ativo=True
            ).first() or Produto.objects.filter(
                nome__icontains='servico', ativo=True
            ).first()

            if produto_servico:
                ItemVenda.objects.create(
                    venda=venda,
                    produto=produto_servico,
                    quantidade=1,
                    preco_unitario=os_obj.valor_final,
                )

            # Marcar OS como entregue
            os_obj.status = 'entregue'
            if not os_obj.data_entrega:
                from django.utils import timezone
                os_obj.data_entrega = timezone.now().date()
            os_obj.save(update_fields=['status', 'data_entrega'])

        messages.success(request, f'Venda #{venda.id} gerada com sucesso! OS marcada como Entregue.')
        return redirect('core:venda_visualizar', venda_id=venda.id)

    # GET: mostrar tela de confirmação
    return render(request, 'core/os_gerar_venda.html', {
        'os': os_obj,
        'forma_pagamento_choices': [
            ('dinheiro', 'Dinheiro'), ('debito', 'Débito'), ('credito', 'Crédito'),
            ('pix', 'PIX'), ('parcelado', 'Parcelado'),
        ],
    })


# ============================================================
#  DEVOLUÇÕES
# ============================================================

@login_required
@verificar_licenca_view
def lista_devolucoes(request):
    """Listar todas as devoluções"""
    qs = Devolucao.objects.select_related('venda', 'produto').order_by('-data_devolucao')
    busca = request.GET.get('busca', '').strip()
    if busca:
        qs = qs.filter(
            Q(produto__nome__icontains=busca) |
            Q(venda__id__icontains=busca) |
            Q(descricao__icontains=busca)
        )
    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))
    total_reembolso = qs.aggregate(s=Sum('valor_reembolso'))['s'] or Decimal('0')
    return render(request, 'core/lista_devolucoes.html', {
        'page_obj': page, 'busca': busca, 'total_reembolso': total_reembolso,
    })


@login_required
@verificar_licenca_view
def nova_devolucao(request, venda_id):
    """Registrar devolução a partir de uma venda"""
    venda = get_object_or_404(Venda, id=venda_id, status='finalizada')
    produtos_da_venda = Produto.objects.filter(
        id__in=venda.items.values_list('produto_id', flat=True)
    )

    form = DevolucaoForm(request.POST or None)
    # Limitar produtos ao que foi vendido
    form.fields['produto'].queryset = produtos_da_venda

    if request.method == 'POST' and form.is_valid():
        dev = form.save(commit=False)
        dev.venda = venda
        dev.registrado_por = request.user.username
        dev.save()
        messages.success(
            request,
            f'Devolução #{dev.id} registrada! '
            f'{"Estoque reposto." if dev.repor_estoque else "Estoque não alterado."}'
        )
        return redirect('core:venda_visualizar', venda_id=venda.id)

    return render(request, 'core/nova_devolucao.html', {
        'form': form, 'venda': venda,
    })


# -----------------------------------------------------------------------
# Views de erro HTTP customizadas
# -----------------------------------------------------------------------

def view_400(request, exception=None):
    """Requisição inválida (400)."""
    return render(request, '400.html', {'exception': exception}, status=400)


def view_403(request, exception=None):
    """Acesso negado (403)."""
    return render(request, '403.html', {'exception': exception}, status=403)


def view_404(request, exception=None):
    """Página não encontrada (404)."""
    return render(request, '404.html', {'exception': exception}, status=404)


def view_500(request):
    """Erro interno do servidor (500) — fallback quando o middleware falha."""
    return render(request, 'core/erro_interno.html', {
        'error_id': 'N/A',
        'error_message': 'Erro interno no servidor.',
        'error_type': 'ServerError',
        'traceback': 'Detalhes não disponíveis. Consulte o arquivo logs/erros.log.',
        'path': request.path,
        'method': request.method,
    }, status=500)
